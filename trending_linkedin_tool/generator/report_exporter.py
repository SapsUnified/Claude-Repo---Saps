"""Export daily reports as Excel (.xlsx) and plain-text (.txt) files."""

import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from trending_linkedin_tool.scrapers.base import ScrapedItem

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CATEGORY_FONT = Font(bold=True, size=11, color="2F5496")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


class ReportExporter:
    """Exports trending topics and post drafts to Excel and text formats."""

    def save_excel(
        self,
        output_dir: str,
        timestamp: str,
        trending_report: dict,
        linkedin_posts: list[dict],
        twitter_posts: list[dict],
    ) -> str:
        """Save a complete daily report as an Excel workbook.

        Returns the path to the saved file.
        """
        wb = Workbook()

        # --- Sheet 1: Trending Topics ---
        ws_trends = wb.active
        ws_trends.title = "Trending Topics"
        self._write_trending_sheet(ws_trends, trending_report, timestamp)

        # --- Sheet 2: LinkedIn Posts ---
        ws_linkedin = wb.create_sheet("LinkedIn Posts")
        self._write_posts_sheet(ws_linkedin, linkedin_posts, "LinkedIn", timestamp)

        # --- Sheet 3: Twitter Posts ---
        ws_twitter = wb.create_sheet("Twitter Posts")
        self._write_posts_sheet(ws_twitter, twitter_posts, "Twitter/X", timestamp)

        path = os.path.join(output_dir, f"daily_report_{timestamp}.xlsx")
        wb.save(path)
        logger.info("Excel report saved to: %s", path)
        return path

    def save_text(
        self,
        output_dir: str,
        timestamp: str,
        trending_report: dict,
        linkedin_posts: list[dict],
        twitter_posts: list[dict],
    ) -> str:
        """Save a complete daily report as a plain-text file.

        Returns the path to the saved file.
        """
        lines: list[str] = []

        lines.append("=" * 60)
        lines.append(f"  DAILY TRENDING TOPICS REPORT — {timestamp}")
        lines.append("=" * 60)
        lines.append("")

        # Trending topics
        for cat_label, items_data in trending_report.items():
            lines.append(f"## {cat_label}")
            lines.append("-" * 40)
            for i, item in enumerate(items_data, 1):
                lines.append(f"  {i}. {item['title']}")
                lines.append(
                    f"     Engagement: {item['engagement']:,} | "
                    f"Source: {item['source']}"
                )
                if item.get("url"):
                    lines.append(f"     URL: {item['url']}")
            lines.append("")

        # LinkedIn posts
        lines.append("=" * 60)
        lines.append("  LINKEDIN POST SUGGESTIONS")
        lines.append("=" * 60)
        lines.append("")

        for post in linkedin_posts:
            lines.append(
                f"--- Post #{post['post_number']} ({post['category']}) ---"
            )
            lines.append(post["full_post"])
            lines.append(f"\nSource: {post['source_url']}")
            lines.append("")

        # Twitter posts
        lines.append("=" * 60)
        lines.append("  TWITTER/X POST SUGGESTIONS")
        lines.append("=" * 60)
        lines.append("")

        for post in twitter_posts:
            lines.append(
                f"--- Tweet #{post['post_number']} ({post['category']}) "
                f"[{post['char_count']} chars] ---"
            )
            lines.append(post["full_post"])
            lines.append(f"\nSource: {post['source_url']}")
            lines.append("")

        path = os.path.join(output_dir, f"daily_report_{timestamp}.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        logger.info("Text report saved to: %s", path)
        return path

    # ---- Internal helpers ----

    def _write_trending_sheet(self, ws, trending_report: dict, timestamp: str):
        """Write trending topics data to an Excel worksheet."""
        # Title row
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"Daily Trending Topics — {timestamp}"
        title_cell.font = Font(bold=True, size=14, color="2F5496")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        row = 3
        for cat_label, items_data in trending_report.items():
            # Category header
            ws.merge_cells(f"A{row}:E{row}")
            cat_cell = ws.cell(row=row, column=1, value=cat_label)
            cat_cell.font = CATEGORY_FONT
            ws.row_dimensions[row].height = 25
            row += 1

            # Column headers
            headers = ["#", "Title", "Source", "Engagement", "URL"]
            col_widths = [5, 50, 15, 15, 50]
            for col, (header, width) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col)].width = width
            row += 1

            # Data rows
            for i, item in enumerate(items_data, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=item["title"])
                ws.cell(row=row, column=3, value=item["source"])
                ws.cell(row=row, column=4, value=item["engagement"])
                url_cell = ws.cell(row=row, column=5, value=item.get("url", ""))
                url_cell.font = Font(color="0563C1", underline="single")

                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                    ws.cell(row=row, column=col).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
                row += 1

            row += 1  # Blank row between categories

    def _write_posts_sheet(
        self, ws, posts: list[dict], platform: str, timestamp: str
    ):
        """Write post drafts to an Excel worksheet."""
        # Title row
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = f"{platform} Post Suggestions — {timestamp}"
        title_cell.font = Font(bold=True, size=14, color="2F5496")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # Column headers
        headers = ["#", "Category", "Post Content", "Source URL"]
        col_widths = [5, 25, 80, 50]
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = width

        # Post rows
        for i, post in enumerate(posts):
            row = 4 + i
            ws.cell(row=row, column=1, value=post["post_number"])
            ws.cell(row=row, column=2, value=post["category"])

            content_cell = ws.cell(row=row, column=3, value=post["full_post"])
            content_cell.alignment = Alignment(wrap_text=True, vertical="top")

            url_cell = ws.cell(row=row, column=4, value=post.get("source_url", ""))
            url_cell.font = Font(color="0563C1", underline="single")

            # Auto-height based on content lines
            line_count = post["full_post"].count("\n") + 1
            ws.row_dimensions[row].height = max(15, line_count * 15)

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER
